# -*- coding: utf-8 -*-
"""
UEVÇB Raporu — EPİAŞ Şeffaflık

Açılınca KGÜP (İlk Versiyon) bölümündeki TÜM organizasyon adlarını listeler
(seçilen tarih aralığının tamamı taranır — aylar içinde eklenen/çıkan firmalar
da görünür). Organizasyon seçilince UEVÇB adları gelir; "Excel Oluştur" ile
GÖP eşleşme + İA alış/satış saatlik verileri UEVÇB adlı Excel dosyalarına yazılır.

NOT: EPİAŞ GÖP eşleşme ve İA verilerini yalnız ORGANİZASYON bazında yayınlar;
UEVÇB kırılımı yoktur. Dosyalar UEVÇB adıyla adlandırılır ama içerik
organizasyon toplamıdır (dosya içine de not düşülür).

Kullanım:
    çift tık: arayuz.bat  (ya da masaüstündeki "UEVCB Raporu" kısayolu)
    komut satırı: python uevcb_rapor.py --org MASAT --bas 2025-01-01 --bit 2026-07-17
"""
import argparse
import calendar
import datetime as dt
import json
import os
import re
import sys
import threading
import time
import queue
from pathlib import Path

import pandas as pd

HERE = Path(__file__).resolve().parent
_E = None  # tek oturumluk EPTR2 bağlantısı


def ayarlari_yukle():
    yol = HERE / "ayarlar.json"
    if not yol.exists():
        raise RuntimeError("ayarlar.json bulunamadı — önce kurulum.bat'ı çalıştırın.")
    try:
        a = json.loads(yol.read_text(encoding="utf-8-sig"))
    except Exception as ex:
        raise RuntimeError(f"ayarlar.json okunamadı: {ex}")
    kul = str(a.get("api_kullanici", "")).strip()
    if not kul or "BURAYA" in kul.upper():
        raise RuntimeError("ayarlar.json: api_kullanici doldurulmamış (EPİAŞ Şeffaflık e-postası).")
    return a


def sifre_bul(ayar):
    pw = os.environ.get("EPIAS_PW")
    if pw:
        return pw.strip()
    pw = str(ayar.get("api_sifre", "")).strip()
    if pw and "BURAYA" not in pw.upper():
        return pw
    dosya = ayar.get("api_sifre_dosyasi", "~/.epias_pw")
    for aday in [Path(dosya).expanduser(), HERE / ".epias_pw"]:
        if aday.exists():
            return aday.read_text().strip()
    raise RuntimeError("EPİAŞ şifresi bulunamadı — ayarlar.json'daki api_sifre alanını doldurun.")


def indirilenler():
    d = Path.home() / "Downloads"
    return d if d.exists() else HERE


def baglan():
    global _E
    if _E is None:
        from eptr2 import EPTR2
        ayar = ayarlari_yukle()
        os.chdir(HERE)  # .eptr2-tgt bileti bu klasöre yazılsın (gitignore'da)
        _E = EPTR2(username=ayar["api_kullanici"].strip(), password=sifre_bul(ayar))
    return _E


def df_yap(res):
    return res if isinstance(res, pd.DataFrame) else pd.DataFrame(res)


def org_listesi(bas, bit):
    """KGÜP organizasyon listesi — aralığın TAMAMI için (birleşim döner,
    ay içinde eklenen/çıkan firmalar da listelenir)."""
    e = baglan()
    df = df_yap(e.call("gen-org", start_date=str(bas), end_date=str(bit)))
    df = df.sort_values("organizationName")
    return list(df[["organizationId", "organizationName"]].itertuples(index=False, name=None))


def ay_baslari(bas, bit):
    b = dt.date.fromisoformat(str(bas)).replace(day=1)
    s = dt.date.fromisoformat(str(bit))
    out = []
    while b <= s:
        out.append(b)
        b = (b + dt.timedelta(days=32)).replace(day=1)
    return out


def uevcb_listesi(org_id, bas, bit, log=lambda m: None):
    """UEVÇB adları — EPİAŞ bu listeyi ay bazında verir; aralıktaki her ay
    taranıp birleşim alınır (ay içinde değişen adlar/santraller kaçmaz)."""
    e = baglan()
    birlesim = {}
    for ay in ay_baslari(bas, bit):
        try:
            df = df_yap(e.call("gen-uevcb", org_id=org_id, start_date=ay.isoformat()))
        except Exception as ex:
            log(f"  uyarı: {ay:%Y-%m} UEVÇB listesi alınamadı ({str(ex)[:60]})")
            continue
        for _, sat in df.iterrows() if len(df) else []:
            birlesim[int(sat["id"])] = str(sat["name"])  # sonraki ay adı değiştiyse yenisi kalır
    return sorted(birlesim.items(), key=lambda x: x[1])


def aylik_dilimler(bas, bit):
    b = dt.date.fromisoformat(str(bas))
    s = dt.date.fromisoformat(str(bit))
    out = []
    while b <= s:
        ay_son = b.replace(day=calendar.monthrange(b.year, b.month)[1])
        out.append((b, min(ay_son, s)))
        b = ay_son + dt.timedelta(days=1)
    return out


def saatlik_cek(key, org_id, bas, bit, kolon, log=print, **ek):
    """Bir endpoint'i aylık parçalarla çeker; tek saatlik seri/sözlük döndürür.
    EPİAŞ hız limitine (429) takılınca bekleyip aynı parçayı yeniden dener."""
    e = baglan()

    def cagri(b, s):
        for deneme in range(6):
            try:
                return df_yap(e.call(key, start_date=b.isoformat(), end_date=s.isoformat(),
                                     org_id=org_id, **ek))
            except Exception as ex:
                if "429" in str(ex):
                    bekle = 4 * (deneme + 1)
                    log(f"  hız limiti (429) — {bekle} sn bekleniyor...")
                    time.sleep(bekle)
                    continue
                raise
        raise RuntimeError("429 hız limiti: yeniden denemeler tükendi")

    parcalar = []
    for b, s in aylik_dilimler(bas, bit):
        time.sleep(0.4)  # hız limitine takılmamak için
        try:
            df = cagri(b, s)
        except Exception as ex:
            # geleceğe taşan uç (yarının verisi 14:00 öncesi yok vb.) — günü geri çekerek dene
            df = None
            s2 = s
            for _ in range(4):
                s2 = s2 - dt.timedelta(days=1)
                if s2 < b:
                    break
                try:
                    df = cagri(b, s2)
                    break
                except RuntimeError:
                    break  # 429 denemeleri tükendi — parçayı atla
                except Exception:
                    continue
            if df is None:
                log(f"  uyarı: {key} {b}—{s} alınamadı, atlandı ({str(ex)[:80]})")
                continue
        if len(df):
            parcalar.append(df)
    if not parcalar:
        return pd.Series(dtype=float) if isinstance(kolon, str) else {k: pd.Series(dtype=float) for k in kolon}
    df = pd.concat(parcalar, ignore_index=True)
    ts = pd.to_datetime(df["date"].str[:19])
    if isinstance(kolon, str):
        return pd.Series(pd.to_numeric(df[kolon], errors="coerce").values, index=ts)
    return {k: pd.Series(pd.to_numeric(df[v], errors="coerce").values, index=ts) for k, v in kolon.items()}


def guvenli_ad(ad):
    return re.sub(r'[\\/:*?"<>|]+', "-", ad).strip()


def rapor_uret(org_id, org_ad, bas, bit, klasor, log=print, uevcbler=None):
    """Organizasyon için TEK Excel üretir: GÖP/İA (org toplamı) + her santralin
    İlk/Son KGÜP'ü + UEVM toplamı + NET (Σ Son−İlk KGÜP, Excel formülü)."""
    from openpyxl.utils import get_column_letter

    log(f"Organizasyon: {org_ad} (id {org_id})")
    if uevcbler is None:
        log("UEVÇB listesi taranıyor (ay ay)...")
        uevcbler = uevcb_listesi(org_id, bas, bit, log)
    if uevcbler:
        log("Santraller: " + ", ".join(ad for _, ad in uevcbler))
    else:
        log("UEVÇB listesi boş — yalnız organizasyon verileri yazılacak.")

    log("GÖP eşleşme çekiliyor...")
    gop = saatlik_cek("dam-clearing", org_id, bas, bit,
                      {"GÖP Eşleşme Alış (MWh)": "matchedBids",
                       "GÖP Eşleşme Satış (MWh)": "matchedOffers"}, log)
    log("İA alış çekiliyor...")
    ia_alis = saatlik_cek("bi-long", org_id, bas, bit, "quantity", log)
    log("İA satış çekiliyor...")
    ia_satis = saatlik_cek("bi-short", org_id, bas, bit, "quantity", log)

    idx = pd.date_range(f"{bas} 00:00", f"{bit} 23:00", freq="h")
    tablo = pd.DataFrame(index=idx)
    tablo["Tarih"] = tablo.index.date
    tablo["Saat"] = tablo.index.strftime("%H:00")
    for ad, seri in list(gop.items()) + [("İA Alış (MWh)", ia_alis), ("İA Satış (MWh)", ia_satis)]:
        tablo[ad] = seri.reindex(idx)

    # her santralin İlk/Son KGÜP'ü ayrı sütun (UEVÇB bazlı yayınlanır)
    n = len(uevcbler)
    for uid, ad in uevcbler:
        log(f"{ad} → KGÜP çekiliyor...")
        ilk = saatlik_cek("kgup-v1", org_id, bas, bit, "toplam", log, region="TR1", uevcb_id=uid)
        son = saatlik_cek("kgup", org_id, bas, bit, "toplam", log, region="TR1", uevcb_id=uid)
        tablo[f"{ad} İlk KGÜP"] = ilk.reindex(idx)
        tablo[f"{ad} Son KGÜP"] = son.reindex(idx)

    # UEVM toplamı — UEVM santral (pp) bazlı yayınlanır; adlar UEVÇB adıyla eşleşir
    uevm_serileri = []
    if uevcbler:
        log("UEVM santral listesi çekiliyor...")
        e = baglan()
        pl = df_yap(e.call("uevm-pp-list"))

        def norm(s):
            # noktalama/boşluk farklarını yok say ("...SANT." ↔ "...SANT" gibi)
            return re.sub(r"[^0-9A-ZÇĞİÖŞÜ]+", " ", str(s).upper()).strip()
        pp_ad = {norm(str(nm).rsplit("-", 1)[0]): pid for pid, nm in zip(pl["id"], pl["name"])}
        for _, ad in uevcbler:
            pid = pp_ad.get(norm(ad))
            if pid is None:
                log(f"  uyarı: '{ad}' UEVM santral listesinde yok — UEVM toplamına katılamadı.")
                continue
            log(f"{ad} → UEVM çekiliyor (pp {pid})...")
            s = saatlik_cek("uevm", org_id, bas, bit, "total", log, pp_id=pid)
            uevm_serileri.append(s.reindex(idx))
    if uevm_serileri:
        tablo["UEVM Toplam (MWh)"] = pd.concat(uevm_serileri, axis=1).sum(axis=1, min_count=1)
    else:
        tablo["UEVM Toplam (MWh)"] = float("nan")

    son_dolu = tablo.iloc[:, 2:].notna().any(axis=1)  # yayınlanmamış kuyruk saatleri at
    if son_dolu.any():
        tablo = tablo.loc[:son_dolu[son_dolu].index[-1]]

    klasor = Path(klasor)
    klasor.mkdir(parents=True, exist_ok=True)
    yol = klasor / f"{guvenli_ad(org_ad)}.xlsx"
    try:
        with pd.ExcelWriter(yol, engine="openpyxl") as w:
            tablo.to_excel(w, index=False, sheet_name="Veri", startrow=1)
            ws = w.sheets["Veri"]
            ws["A1"] = (f"{org_ad} — GÖP/İA org toplamı (EPİAŞ UEVÇB kırılımı yayınlamaz); "
                        "KGÜP sütunları santral bazlı; UEVM ~1,5 ay geriden yayınlanır; "
                        "NET = Σ(Son KGÜP − İlk KGÜP).")
            # NET sütunu: santrallerin (Son − İlk KGÜP) toplamı, Excel formülü
            if n:
                net_kol = 8 + 2 * n  # A..F=6, KGÜP çiftleri G'den, +1 UEVM, +1 NET
                ws.cell(row=2, column=net_kol, value="NET (MWh)")
                for r in range(3, len(tablo) + 3):
                    parcalar = [f"({get_column_letter(8 + 2 * i)}{r}-{get_column_letter(7 + 2 * i)}{r})"
                                for i in range(n)]
                    ws.cell(row=r, column=net_kol, value="=" + "+".join(parcalar))
                ws.column_dimensions[get_column_letter(net_kol)].width = 12
            for k, gen in zip("ABCDEF", (12, 7, 22, 22, 16, 16)):
                ws.column_dimensions[k].width = gen
            for i in range(2 * n + 1):  # KGÜP sütunları + UEVM
                ws.column_dimensions[get_column_letter(7 + i)].width = 16
            ws.freeze_panes = "A3"
    except PermissionError:
        log(f"UYARI: {yol.name} yazılamadı — dosya Excel'de açık, kapatıp yeniden deneyin.")
        return []
    log(f"Yazıldı: {yol}")
    log("TAMAM.")
    return [yol]


# ----------------------------- arayüz -----------------------------

def arayuz():
    import tkinter as tk
    from tkinter import ttk, messagebox

    pen = tk.Tk()
    pen.title("UEVÇB Raporu — EPİAŞ (KGÜP organizasyonları)")
    pen.geometry("760x640")
    cer = ttk.Frame(pen, padding=10)
    cer.pack(fill="both", expand=True)

    kuyruk = queue.Queue()
    orglar = []        # [(id, ad)] — tam liste
    suzgun = []        # filtre uygulanmış görünüm
    uevcbler = []      # seçili org'un UEVÇB'leri
    son_org = [None]   # en son seçilen organizasyon (liste yenilense de hatırlanır)
    uevcb_orgu = [None]  # uevcbler listesi hangi org için çekildi
    mesgul = [False]

    # --- üst satır: tarih aralığı + klasör
    ust = ttk.Frame(cer)
    ust.pack(fill="x")
    ttk.Label(ust, text="Başlangıç:").pack(side="left")
    bas_g = ttk.Entry(ust, width=11)
    bas_g.pack(side="left", padx=(2, 10))
    bas_g.insert(0, "2025-01-01")
    ttk.Label(ust, text="Bitiş:").pack(side="left")
    bit_g = ttk.Entry(ust, width=11)
    bit_g.pack(side="left", padx=(2, 10))
    bit_g.insert(0, dt.date.today().isoformat())
    ttk.Label(ust, text="Kayıt klasörü:").pack(side="left")
    klasor_g = ttk.Entry(ust)
    klasor_g.pack(side="left", fill="x", expand=True, padx=2)
    klasor_g.insert(0, str(indirilenler()))

    # --- orta: solda organizasyonlar, sağda UEVÇB'ler
    orta = ttk.Frame(cer)
    orta.pack(fill="both", expand=True, pady=8)
    sol = ttk.LabelFrame(orta, text="Organizasyonlar (KGÜP — İlk Versiyon)", padding=4)
    sol.pack(side="left", fill="both", expand=True, padx=(0, 6))
    sag = ttk.LabelFrame(orta, text="UEVÇB'ler", padding=4)
    sag.pack(side="left", fill="both", expand=True)

    filt_sat = ttk.Frame(sol)
    filt_sat.pack(fill="x")
    ttk.Label(filt_sat, text="Ara:").pack(side="left")
    filtre_g = ttk.Entry(filt_sat)
    filtre_g.pack(side="left", fill="x", expand=True, padx=2)
    yenile_b = ttk.Button(filt_sat, text="Listeyi Yenile")
    yenile_b.pack(side="left")

    org_lb = tk.Listbox(sol, exportselection=False)
    org_sc = ttk.Scrollbar(sol, orient="vertical", command=org_lb.yview)
    org_lb.configure(yscrollcommand=org_sc.set)
    org_lb.pack(side="left", fill="both", expand=True)
    org_sc.pack(side="left", fill="y")

    uevcb_lb = tk.Listbox(sag, exportselection=False, selectmode="extended")
    uevcb_sc = ttk.Scrollbar(sag, orient="vertical", command=uevcb_lb.yview)
    uevcb_lb.configure(yscrollcommand=uevcb_sc.set)
    uevcb_lb.pack(side="left", fill="both", expand=True)
    uevcb_sc.pack(side="left", fill="y")

    # --- alt: butonlar + log
    alt = ttk.Frame(cer)
    alt.pack(fill="x")
    uevcb_b = ttk.Button(alt, text="UEVÇB'leri Göster")
    uevcb_b.pack(side="left")
    excel_b = ttk.Button(alt, text="Excel Oluştur")
    excel_b.pack(side="left", padx=6)
    durum = ttk.Label(alt, text="")
    durum.pack(side="left", padx=8)

    cikti = tk.Text(cer, height=8, state="disabled")
    cikti.pack(fill="both", pady=(6, 0))

    def log(msg):
        kuyruk.put(str(msg))

    def tarihler():
        return bas_g.get().strip(), bit_g.get().strip()

    def calistir_arka(f):
        if mesgul[0]:
            messagebox.showinfo("Bekle", "Devam eden bir işlem var.")
            return
        mesgul[0] = True
        durum.config(text="çalışıyor...")

        def sarici():
            try:
                f()
            except Exception as ex:
                log(f"HATA: {ex}")
            finally:
                mesgul[0] = False
                kuyruk.put("__DURUM_BOS__")
        threading.Thread(target=sarici, daemon=True).start()

    def org_goster():
        org_lb.delete(0, "end")
        suzgun.clear()
        f = filtre_g.get().strip().upper()
        for oid, ad in orglar:
            if not f or f in ad.upper():
                suzgun.append((oid, ad))
                org_lb.insert("end", ad)

    def listeyi_yenile():
        bas, bit = tarihler()

        def is_():
            log(f"Organizasyon listesi çekiliyor ({bas} → {bit}, tüm aylar dahil)...")
            orglar.clear()
            orglar.extend(org_listesi(bas, bit))
            log(f"{len(orglar)} organizasyon listelendi.")
            kuyruk.put("__ORG_GOSTER__")
        calistir_arka(is_)

    def secili_org():
        s = org_lb.curselection()
        if s and suzgun:
            son_org[0] = suzgun[s[0]]
        return son_org[0]  # liste yenilenip seçim kaybolsa da son seçim geçerli

    def uevcb_goster():
        sec = secili_org()
        if not sec:
            messagebox.showwarning("Uyarı", "Soldan bir organizasyon seçin.")
            return
        oid, ad = sec
        bas, bit = tarihler()

        def is_():
            log(f"{ad} → UEVÇB listesi taranıyor ({bas} → {bit}, ay ay)...")
            uevcbler.clear()
            uevcbler.extend(uevcb_listesi(oid, bas, bit, log))
            uevcb_orgu[0] = oid
            log(f"{len(uevcbler)} UEVÇB bulundu — istersen sağdan seç (seçmezsen hepsi Excel'e girer).")
            kuyruk.put("__UEVCB_GOSTER__")
        calistir_arka(is_)

    def excel_olustur():
        sec = secili_org()
        if not sec:
            messagebox.showwarning("Uyarı", "Soldan bir organizasyon seçin "
                                   "(ya da çift tıklayıp UEVÇB'lerini getirin).")
            return
        oid, ad = sec
        bas, bit = tarihler()
        klasor = klasor_g.get().strip()
        u = None
        if uevcbler and uevcb_orgu[0] == oid:  # başka org'un UEVÇB listesi kullanılmasın
            secili = [uevcbler[i] for i in uevcb_lb.curselection()]
            if secili and len(secili) < len(uevcbler):
                # farkında olmadan eksik santralli dosya üretilmesin — açıkça sor
                if not messagebox.askyesno(
                        "Santral seçimi",
                        f"Sağda {len(secili)}/{len(uevcbler)} santral seçili:\n"
                        + "\n".join("• " + ad for _, ad in secili)
                        + "\n\nYalnız seçilenler dosyaya girecek. Devam edilsin mi?\n"
                        "(Hayır = TÜM santraller dahil edilir)"):
                    secili = []
            u = secili or list(uevcbler)  # sağda seçim varsa yalnız onlar, yoksa hepsi

        def is_():
            rapor_uret(oid, ad, bas, bit, klasor, log=log, uevcbler=u)
        calistir_arka(is_)

    def kuyruk_isle2():  # widget güncellemeleri ana thread'de
        try:
            while True:
                m = kuyruk.get_nowait()
                if m == "__ORG_GOSTER__":
                    org_goster()
                elif m == "__UEVCB_GOSTER__":
                    uevcb_lb.delete(0, "end")
                    for _, ad in uevcbler:
                        uevcb_lb.insert("end", ad)
                elif m == "__DURUM_BOS__":
                    durum.config(text="")
                else:
                    cikti.configure(state="normal")
                    cikti.insert("end", m + "\n")
                    cikti.see("end")
                    cikti.configure(state="disabled")
        except queue.Empty:
            pass
        pen.after(150, kuyruk_isle2)

    yenile_b.configure(command=listeyi_yenile)
    uevcb_b.configure(command=uevcb_goster)
    excel_b.configure(command=excel_olustur)
    filtre_g.bind("<KeyRelease>", lambda e: org_goster())
    org_lb.bind("<<ListboxSelect>>", lambda e: secili_org())  # tık anında seçimi hatırla
    org_lb.bind("<Double-Button-1>", lambda e: uevcb_goster())

    pen.after(150, kuyruk_isle2)
    pen.after(300, listeyi_yenile)  # açılışta tüm organizasyonlar otomatik gelsin
    pen.mainloop()


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--org", help="organizasyon adı (arama); verilmezse arayüz açılır")
    p.add_argument("--bas", default="2025-01-01")
    p.add_argument("--bit", default=dt.date.today().isoformat())
    p.add_argument("--klasor", default=str(indirilenler()))
    a = p.parse_args()
    if not a.org:
        arayuz()
        return
    adaylar = [(o, ad) for o, ad in org_listesi(a.bas, a.bit) if a.org.upper() in ad.upper()]
    if not adaylar:
        sys.exit(f"HATA: '{a.org}' adında organizasyon bulunamadı.")
    if len(adaylar) > 1:
        print("Birden çok eşleşme, ilki kullanılıyor:")
        for oid, ad in adaylar:
            print(f"  {oid}  {ad}")
    oid, ad = adaylar[0]
    rapor_uret(oid, ad, a.bas, a.bit, a.klasor)


if __name__ == "__main__":
    main()
